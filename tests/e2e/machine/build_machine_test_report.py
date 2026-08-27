"""Builds the combined Unit + Playwright E2E Excel report for the Machine
module from the current test run's JSON outputs.

Inputs (must already exist — run the suites first):
  - test-report/results/results.json      (Playwright JSON reporter, per playwright.config.js)
  - test-report/results/unit-results.json (Vitest JSON reporter: `vitest run --reporter=json --outputFile=...`)

Output:
  - reports/machine-test-report.xlsx, with 3 sheets: Summary, Unit Tests, E2E Tests.

Generalized from an earlier version that only matched "MCH-###" titles — this
one recognizes every test-ID prefix actually in use across the Machine specs
(MCH-, MCHX-, MCHB-, SMK-), and prefers an explicit "(positive)"/"(negative)"
marker in the title (used by machine-buttons.spec.js) over keyword-guessing
when one is present, falling back to the keyword heuristic otherwise.
"""
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
E2E_RESULTS = ROOT / 'test-report' / 'results' / 'results.json'
UNIT_RESULTS = ROOT / 'test-report' / 'results' / 'unit-results.json'
OUTPUT = ROOT / 'reports' / 'machine-test-report.xlsx'

STATUS_MAP = {'passed': 'Passed', 'failed': 'Failed', 'skipped': 'Skipped', 'timedOut': 'Failed', 'interrupted': 'Failed'}
ANSI_RE = re.compile(r'\x1b\[[0-9;]*m')
# Matches "MCH-004 - ...", "MCHX-115 - ...", "MCHB-01 (positive) - ...", "SMK-01 - ...".
TITLE_RE = re.compile(r'^([A-Z]+-\d+)\s*(?:\(([a-zA-Z]+)\))?\s*-\s*(.*)$')
NEGATIVE_KEYWORDS = ('failure', 'invalid', 'error', 'empty', 'blocks', 'blocked', 'not ', 'never ', 'does not', 'stays hidden', 'without', 'no false', 'race', 'malformed')


def iter_e2e_specs(suites):
    for suite in suites:
        yield from suite.get('specs', [])
        yield from iter_e2e_specs(suite.get('suites', []))


def classify_kind(marker, title):
    if marker:
        m = marker.lower()
        if m.startswith('pos'):
            return 'Positive'
        if m.startswith('neg'):
            return 'Negative'
    lowered = title.lower()
    return 'Negative' if any(word in lowered for word in NEGATIVE_KEYWORDS) else 'Positive'


def build_e2e_rows():
    if not E2E_RESULTS.exists():
        return []
    data = json.loads(E2E_RESULTS.read_text(encoding='utf-8'))
    rows = []
    for spec in iter_e2e_specs(data.get('suites', [])):
        match = TITLE_RE.match(spec['title'])
        if not match:
            continue
        test_id, marker, title = match.group(1), match.group(2), match.group(3)
        test_entry = spec['tests'][0]
        result = test_entry['results'][0]
        state = STATUS_MAP.get(result['status'], 'Skipped')
        error = ''
        screenshot = ''
        if result.get('errors'):
            error = ANSI_RE.sub('', result['errors'][0].get('message', '')).split('\n')[0]
        for attachment in result.get('attachments', []):
            if attachment.get('name') == 'screenshot':
                screenshot = attachment.get('path', '')
        kind = classify_kind(marker, title)
        duration_ms = result.get('duration', 0)
        rows.append({
            'id': test_id, 'title': title, 'kind': kind, 'status': state,
            'duration_ms': duration_ms, 'error': error, 'screenshot': screenshot,
            'retries': len(test_entry['results']) - 1,
        })
    # Stable order: numeric within each prefix, prefixes in first-seen order.
    prefix_order = []
    for r in rows:
        p = r['id'].split('-')[0]
        if p not in prefix_order:
            prefix_order.append(p)

    def sort_key(r):
        p, n = r['id'].split('-')
        return (prefix_order.index(p), int(n))
    rows.sort(key=sort_key)
    return rows


UNIT_MARKER_RE = re.compile(r'\b(positive|negative|boundary)\s*:', re.IGNORECASE)


def build_unit_rows():
    if not UNIT_RESULTS.exists():
        return []
    data = json.loads(UNIT_RESULTS.read_text(encoding='utf-8'))
    rows = []
    for file_result in data.get('testResults', []):
        file_name = Path(file_result.get('name', '')).name
        for assertion in file_result.get('assertionResults', []):
            status = assertion.get('status', 'skipped')
            state = 'Passed' if status == 'passed' else ('Failed' if status == 'failed' else 'Skipped')
            title = assertion.get('fullName') or assertion.get('title', '')
            failure_msgs = assertion.get('failureMessages') or []
            error = ANSI_RE.sub('', failure_msgs[0]).split('\n')[0] if failure_msgs else ''
            duration_ms = assertion.get('duration') or 0
            # This suite's own `it('positive: ...')` / `it('negative: ...')`
            # / `it('boundary: ...')` convention (see any tests/unit/*.test.js)
            # is authoritative when present — read it directly rather than
            # guess from keywords. Boundary cases test an edge of VALID
            # input (e.g. exactly 0, exactly the max) and are their own
            # category, not folded into Negative (which is reserved for
            # actually invalid/malformed input).
            marker = UNIT_MARKER_RE.search(title)
            if marker:
                kind = marker.group(1).capitalize()
            else:
                lowered = title.lower()
                kind = 'Negative' if any(word in lowered for word in ('invalid', 'null', 'malformed', 'nan', 'never', 'not ', 'duplicate')) else 'Positive'
            rows.append({
                'file': file_name, 'title': title, 'kind': kind, 'status': state,
                'duration_ms': duration_ms, 'error': error,
            })
    return rows


def style_header(ws, row_idx=1):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F3D2E')


def autosize(ws, widths):
    for col, width in zip('ABCDEFGHIJ', widths):
        ws.column_dimensions[col].width = width


def status_fill(status):
    return {'Passed': 'E6F6EE', 'Failed': 'FDEAEA', 'Skipped': 'FBF2E2'}.get(status, 'FFFFFF')


def main():
    e2e_rows = build_e2e_rows()
    unit_rows = build_unit_rows()

    wb = Workbook()

    # --- Summary ------------------------------------------------------------
    summary = wb.active
    summary.title = 'Overall Summary'
    summary.append(['Electric Idli Machine — Unit + Playwright E2E Test Report'])
    summary['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    summary['A1'].fill = PatternFill('solid', fgColor='1F3D2E')
    summary.merge_cells('A1:C1')
    summary.append([])

    def block(title, rows, id_key='id'):
        total = len(rows)
        passed = sum(r['status'] == 'Passed' for r in rows)
        failed = sum(r['status'] == 'Failed' for r in rows)
        skipped = sum(r['status'] == 'Skipped' for r in rows)
        positive = sum(r['kind'] == 'Positive' for r in rows)
        negative = sum(r['kind'] == 'Negative' for r in rows)
        boundary = sum(r['kind'] == 'Boundary' for r in rows)
        summary.append([title])
        summary['A' + str(summary.max_row)].font = Font(size=12, bold=True)
        stat_lines = [
            ('Total test cases', total), ('Passed', passed), ('Failed', failed), ('Skipped', skipped),
            ('Positive cases', positive), ('Negative cases', negative),
        ]
        if boundary:
            stat_lines.append(('Boundary cases', boundary))
        stat_lines.append(('Pass percentage', f'{(passed / total * 100):.1f}%' if total else '0.0%'))
        for label, value in stat_lines:
            summary.append(['', label, value])
        summary.append([])

    block('Unit Tests (Vitest)', unit_rows)
    block('Playwright E2E Tests (Machine module)', e2e_rows)

    summary.append(['Run sources'])
    summary.append(['', 'Unit results', str(UNIT_RESULTS)])
    summary.append(['', 'E2E results', str(E2E_RESULTS)])
    summary.column_dimensions['A'].width = 24
    summary.column_dimensions['B'].width = 26
    summary.column_dimensions['C'].width = 90

    # --- Unit Tests sheet -----------------------------------------------------
    ws_unit = wb.create_sheet('Unit Tests')
    headers = ['File', 'Test Case', 'Type', 'Status', 'Duration (ms)', 'Error']
    ws_unit.append(headers)
    style_header(ws_unit)
    for r in unit_rows:
        ws_unit.append([r['file'], r['title'], r['kind'], r['status'], r['duration_ms'], r['error']])
    for row in ws_unit.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        row[3].fill = PatternFill('solid', fgColor=status_fill(row[3].value))
    autosize(ws_unit, [26, 90, 12, 12, 14, 70])
    ws_unit.freeze_panes = 'A2'

    # --- E2E Tests sheet --------------------------------------------------
    ws_e2e = wb.create_sheet('E2E Tests')
    headers = ['Test ID', 'Test Name', 'Category', 'Positive/Negative', 'Preconditions', 'Steps', 'Expected Result', 'Actual Result', 'Pass/Fail', 'Error Message', 'Screenshot/trace/video path', 'API status']
    ws_e2e.append(headers)
    style_header(ws_e2e)

    def category_for(r):
        title = r['title'].lower()
        if 'settings' in title or 'threshold' in title or 'buzzer' in title or 'sync' in title:
            return 'Settings / Controls'
        if 'heater' in title or 'machine runtime' in title or 'temperature' in title or 'chart' in title:
            return 'Telemetry / Controls'
        if 'history' in title or 'range' in title or 'command' in title or 'event' in title or 'graph' in title:
            return 'History / Search / Filter'
        return 'Navigation / Stability'

    def api_status_for(r):
        title = r['title'].lower()
        for code in ('404', '500'):
            if code in title:
                return code + ' (intercepted error path)'
        if 'network failure' in title or 'malformed' in title:
            return 'network/malformed response (intercepted)'
        if 'patch' in title or 'sync' in title:
            return 'PATCH/POST 200 (mocked)'
        return 'GET 200 (mocked where API behavior is asserted)'

    def append_e2e_row(ws, r):
        expected = 'All assertions in the automated scenario pass.'
        actual = 'All assertions passed.' if r['status'] == 'Passed' else (r['error'] or 'Automated assertion failed.')
        ws.append([
            r['id'], r['title'], category_for(r), r['kind'],
            'Authenticated administrator; deterministic machine API fixture where required.',
            'Run Playwright scenario: ' + r['title'], expected, actual, r['status'],
            r['error'], r['screenshot'], api_status_for(r),
        ])

    for r in e2e_rows:
        append_e2e_row(ws_e2e, r)
    for row in ws_e2e.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        row[8].fill = PatternFill('solid', fgColor=status_fill(row[8].value))
    autosize(ws_e2e, [12, 55, 24, 18, 44, 58, 45, 42, 12, 65])
    ws_e2e.freeze_panes = 'A2'

    # Requested drill-downs use the same complete evidence columns rather
    # than a lossy summary, so a reviewer can inspect only positive,
    # negative, or failed scenarios without filtering the main sheet.
    for title, rows in (
        ('Positive Tests', [r for r in e2e_rows if r['kind'] == 'Positive']),
        ('Negative Tests', [r for r in e2e_rows if r['kind'] == 'Negative']),
        ('Failed Tests - Defects', [r for r in e2e_rows if r['status'] == 'Failed']),
    ):
        ws = wb.create_sheet(title)
        ws.append(headers)
        style_header(ws)
        for r in rows:
            append_e2e_row(ws, r)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            row[8].fill = PatternFill('solid', fgColor=status_fill(row[8].value))
        autosize(ws, [12, 55, 24, 18, 44, 58, 45, 42, 12, 65])
        ws.freeze_panes = 'A2'

    OUTPUT.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT)

    u_total, u_pass = len(unit_rows), sum(r['status'] == 'Passed' for r in unit_rows)
    e_total, e_pass = len(e2e_rows), sum(r['status'] == 'Passed' for r in e2e_rows)
    print(f'Saved {OUTPUT}')
    print(f'Unit: {u_pass}/{u_total} passed')
    print(f'E2E:  {e_pass}/{e_total} passed')


if __name__ == '__main__':
    main()
