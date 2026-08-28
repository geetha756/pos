"""
Builds reports/full-e2e-test-report.xlsx — the complete, consolidated
end-to-end Playwright test report across BOTH the Staff Management and
Electric Idli Machine modules, covering every spec file under tests/e2e/
(TC-*, NEG-*, NEW-STF-*, MCH-*, MCHX-*, MCHP-*, SMK-*, NEW-MCH-*, etc.).

Input: test-report/results/results.json (Playwright JSON reporter, written
automatically by playwright.config.js on a plain `npx playwright test` run
with no --reporter override).

    npx playwright test tests/e2e/
    python tests/e2e/build_full_e2e_report.py

Workbook layout: Summary (overall + per-module + per-prefix breakdown),
then "Staff Test Cases" and "Machine Test Cases" detail sheets, then
"Failed Tests" for a quick defect list.
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent.parent
RESULTS_JSON = PROJECT_ROOT / 'test-report' / 'results' / 'results.json'
OUT_XLSX = PROJECT_ROOT / 'reports' / 'full-e2e-test-report.xlsx'

if not RESULTS_JSON.exists():
    print(f'No results file at {RESULTS_JSON} - run `npx playwright test tests/e2e/` first.', file=sys.stderr)
    sys.exit(1)

OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

FONT = 'Arial'
NAVY = '1F3D2E'
TEAL = '157A5E'
PASS_GREEN = '1F9D63'
PASS_SOFT = 'E6F6EE'
FAIL_RED = 'D64545'
FAIL_SOFT = 'FDEAEA'
SKIP_AMBER = 'B98A2E'
SKIP_SOFT = 'FBF2E2'
NEG_PLUM = '7A3B69'
NEG_SOFT = 'F3E7F0'
POS_TEAL_SOFT = 'E7F3EE'
GREY = '5C6B63'
BORDER_CLR = 'DDE2DC'
thin = Side(style='thin', color=BORDER_CLR)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

ANSI_RE = re.compile(r'\x1b\[[0-9;]*m')
# Matches "TC-001: ...", "NEG-011: ...", "MCH-001 - ...", "MCHX-115 - ...",
# "SMK-01 - ...", "NEW-STF-013 (positive) - ...", "NEW-MCH-003 (negative) - ...".
# The ID group is everything up to the first ":" or " - "/"- " separator,
# so multi-hyphen prefixes like NEW-STF- are captured whole.
TITLE_RE = re.compile(r'^([A-Za-z0-9]+(?:-[A-Za-z0-9]+)*)\s*(?:\(([a-zA-Z]+)\))?\s*[:\-]\s*(.*)$')
STATUS_MAP = {'passed': 'PASS', 'failed': 'FAIL', 'timedOut': 'FAIL', 'interrupted': 'FAIL', 'skipped': 'SKIPPED'}
NEGATIVE_KEYWORDS = (
    'failure', 'invalid', 'error', 'empty', 'blocks', 'blocked', 'not ', 'never ',
    'does not', "doesn't", 'stays hidden', 'without', 'no false', 'malformed',
    'absent', 'rejected', 'redirects to login', 'non-existent', 'left blank',
)


def walk_specs(suites, file_name=None):
    for suite in suites:
        fname = suite.get('file') or file_name
        for spec in suite.get('specs', []):
            yield fname, spec
        yield from walk_specs(suite.get('suites', []), fname)


def classify_kind(marker, scenario):
    if marker:
        m = marker.lower()
        if m.startswith('pos'):
            return 'Positive'
        if m.startswith('neg'):
            return 'Negative'
    lowered = scenario.lower()
    return 'Negative' if any(word in lowered for word in NEGATIVE_KEYWORDS) else 'Positive'


data = json.loads(RESULTS_JSON.read_text(encoding='utf-8'))

rows = []
unmatched_titles = []
for fname, spec in walk_specs(data['suites']):
    match = TITLE_RE.match(spec['title'])
    if not match:
        unmatched_titles.append(spec['title'])
        continue
    tc_id, marker, scenario = match.group(1), match.group(2), match.group(3)
    category = classify_kind(marker, scenario)

    test = spec['tests'][0]
    result = test['results'][0] if test['results'] else {}
    status = STATUS_MAP.get(result.get('status'), 'SKIPPED')
    duration_s = round((result.get('duration') or 0) / 1000, 1)

    error_message = ''
    if status == 'FAIL':
        errors = result.get('errors') or []
        msg = errors[0].get('message', '') if errors else 'Test failed.'
        error_message = ANSI_RE.sub('', msg).strip()

    if status == 'PASS':
        actual = 'All assertions passed as expected.'
    elif status == 'FAIL':
        actual = (error_message.split('\n')[0] or 'Test failed.')[:500]
    else:
        actual = 'Not executed (skipped - typically a cascade skip from an earlier failure in the same serial block).'

    module = 'Machine' if ('machine' in (fname or '').lower()) else 'Staff'
    prefix = re.match(r'^[A-Za-z]+', tc_id).group(0)
    rows.append({
        'id': tc_id, 'prefix': prefix, 'module': module, 'category': category,
        'scenario': scenario, 'file': fname, 'actual': actual, 'status': status,
        'error_message': error_message, 'duration': f'{duration_s}s',
    })


def sort_key(r):
    m = re.search(r'(\d+)$', r['id'])
    return (r['module'], r['prefix'], int(m.group(1)) if m else 0)


rows.sort(key=sort_key)

staff_rows = [r for r in rows if r['module'] == 'Staff']
machine_rows = [r for r in rows if r['module'] == 'Machine']
failed_rows = [r for r in rows if r['status'] == 'FAIL']

status_style = {
    'PASS': (PASS_GREEN, PASS_SOFT),
    'FAIL': (FAIL_RED, FAIL_SOFT),
    'SKIPPED': (SKIP_AMBER, SKIP_SOFT),
}
type_style = {
    'Positive': (TEAL, POS_TEAL_SOFT),
    'Negative': (NEG_PLUM, NEG_SOFT),
}
HEADERS = ['Test Case ID', 'Test Type', 'Spec File', 'Scenario', 'Actual Result', 'Status', 'Error Message', 'Execution Time']
WIDTHS = {'A': 16, 'B': 11, 'C': 30, 'D': 55, 'E': 36, 'F': 10, 'G': 40, 'H': 13}


def write_sheet(ws, sheet_rows, table_name):
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = border_all

    for r, row in enumerate(sheet_rows, start=2):
        values = [row['id'], row['category'], (row['file'] or '').replace('\\', '/'), row['scenario'],
                  row['actual'], row['status'], row['error_message'], row['duration']]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=9.5)
            cell.border = border_all
            cell.alignment = Alignment(vertical='top', wrap_text=True)

        type_cell = ws.cell(row=r, column=2)
        tfg, tbg = type_style.get(row['category'], (GREY, 'FFFFFF'))
        type_cell.font = Font(name=FONT, size=9.5, bold=True, color=tfg)
        type_cell.fill = PatternFill('solid', fgColor=tbg)
        type_cell.alignment = Alignment(horizontal='center', vertical='center')

        status_cell = ws.cell(row=r, column=6)
        fg, bg = status_style.get(row['status'], (GREY, 'FFFFFF'))
        status_cell.font = Font(name=FONT, size=10, bold=True, color=fg)
        status_cell.fill = PatternFill('solid', fgColor=bg)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=r, column=1).font = Font(name=FONT, size=9.5, bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center', vertical='top')
        ws.row_dimensions[r].height = 40

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    last_row = len(sheet_rows) + 1
    if sheet_rows:
        tab = Table(displayName=table_name, ref=f'A1:H{last_row}')
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showFirstColumn=False)
        ws.add_table(tab)


wb = Workbook()
wb.remove(wb.active)

ws_staff = wb.create_sheet('Staff Test Cases')
write_sheet(ws_staff, staff_rows, 'StaffTestCases')

ws_machine = wb.create_sheet('Machine Test Cases')
write_sheet(ws_machine, machine_rows, 'MachineTestCases')

ws_failed = wb.create_sheet('Failed Tests')
write_sheet(ws_failed, failed_rows, 'FailedTestCases')

# ---------- Summary sheet (first tab) ----------
ws2 = wb.create_sheet('Summary', 0)
ws2.sheet_view.showGridLines = False

ws2['B2'] = 'End-to-End Test Report — Staff Management & Electric Idli Machine'
ws2['B2'].font = Font(name=FONT, size=16, bold=True, color=NAVY)
ws2['B3'] = 'Playwright · Chromium · Full regression suite (all specs under tests/e2e/)'
ws2['B3'].font = Font(name=FONT, size=11, italic=True, color=GREY)
ws2['B4'] = f'Run date: {data["stats"].get("startTime", "")[:10]}'
ws2['B4'].font = Font(name=FONT, size=10, color=GREY)


def count(sheet_rows, status):
    return sum(1 for r in sheet_rows if r['status'] == status)


row0 = 6
sections = [
    ('All Test Cases', rows),
    ('Staff Management', staff_rows),
    ('Electric Idli Machine', machine_rows),
]
r = row0
for label, sheet_rows in sections:
    ws2.cell(row=r, column=2, value=label).font = Font(name=FONT, size=11, bold=True, color=NAVY)
    r += 1
    for i, h in enumerate(['Total', 'Passed', 'Failed', 'Skipped', 'Positive', 'Negative'], start=2):
        c = ws2.cell(row=r, column=i, value=h)
        c.font = Font(name=FONT, size=9.5, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center')
    r += 1
    positive_n = sum(1 for x in sheet_rows if x['category'] == 'Positive')
    negative_n = sum(1 for x in sheet_rows if x['category'] == 'Negative')
    values = [len(sheet_rows), count(sheet_rows, 'PASS'), count(sheet_rows, 'FAIL'), count(sheet_rows, 'SKIPPED'), positive_n, negative_n]
    fills = [None, PASS_SOFT, FAIL_SOFT, SKIP_SOFT, POS_TEAL_SOFT, NEG_SOFT]
    fonts = [GREY, PASS_GREEN, FAIL_RED, SKIP_AMBER, TEAL, NEG_PLUM]
    for i, (val, fill, color) in enumerate(zip(values, fills, fonts), start=2):
        c = ws2.cell(row=r, column=i, value=val)
        c.font = Font(name=FONT, size=12, bold=True, color=color)
        c.alignment = Alignment(horizontal='center')
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
    r += 2

overall_row = r
total_failed = count(rows, 'FAIL')
total = len(rows)
total_passed = count(rows, 'PASS')
overall = 'FAIL' if total_failed > 0 else ('PASS' if total_passed == total else 'PARTIAL')
ws2.cell(row=overall_row, column=2, value='Overall Result').font = Font(name=FONT, size=11, bold=True, color=GREY)
oc = ws2.cell(row=overall_row, column=3, value=overall)
oc.font = Font(name=FONT, size=14, bold=True, color=TEAL if overall == 'PASS' else FAIL_RED)
ws2.cell(row=overall_row, column=2).fill = PatternFill('solid', fgColor='E7F3EE')
ws2.cell(row=overall_row, column=3).fill = PatternFill('solid', fgColor='E7F3EE')

note_row = overall_row + 2
ws2.cell(row=note_row, column=2, value='Sheets:').font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws2.cell(row=note_row + 1, column=2,
         value='"Staff Test Cases" and "Machine Test Cases" hold the full per-test detail (scenario, actual result, status, error message, execution time, spec file). "Failed Tests" lists only failures across both modules for quick triage. This is a static export from one Playwright run — re-run the full suite and this script to refresh it.').font = Font(name=FONT, size=10, color=GREY)
ws2.cell(row=note_row + 1, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws2.merge_cells(start_row=note_row + 1, start_column=2, end_row=note_row + 3, end_column=8)

for col, width in zip('ABCDEFGH', [3, 24, 12, 10, 10, 10, 10, 10]):
    ws2.column_dimensions[col].width = width

wb.save(str(OUT_XLSX))
print('Saved:', OUT_XLSX)
print(f'Staff:   {len(staff_rows)} (PASS={count(staff_rows, "PASS")} FAIL={count(staff_rows, "FAIL")} SKIPPED={count(staff_rows, "SKIPPED")})')
print(f'Machine: {len(machine_rows)} (PASS={count(machine_rows, "PASS")} FAIL={count(machine_rows, "FAIL")} SKIPPED={count(machine_rows, "SKIPPED")})')
if unmatched_titles:
    print(f'Warning: {len(unmatched_titles)} test titles did not match the ID pattern and were skipped:', file=sys.stderr)
    for t in unmatched_titles[:20]:
        print('  -', t, file=sys.stderr)
