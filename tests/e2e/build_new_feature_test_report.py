"""
Builds reports/new-feature-test-report.xlsx — a combined Positive/Negative
Excel report for this session's new coverage across BOTH modules:

  tests/e2e/07-staff-validation-positive-negative.spec.js       (NEW-STF-*)
  tests/e2e/machine/machine-timestamp-positive-negative.spec.js (NEW-MCH-*)

Input: test-report/results/results.json (Playwright JSON reporter). Run the
two spec files first, then this script:

    npx playwright test tests/e2e/07-staff-validation-positive-negative.spec.js \\
        tests/e2e/machine/machine-timestamp-positive-negative.spec.js --reporter=json > test-report/results/results.json
    python tests/e2e/build_new_feature_test_report.py

Test titles are self-describing (`NEW-STF-001 (positive) - <scenario>`), so
this reads Positive/Negative and the scenario text straight from each
title - no separate catalog file, matching machine/build_machine_test_report.py's
approach rather than staff's catalog-driven one, since there's no existing
catalog entry for these new IDs to reuse.

Workbook layout: Summary, then one sheet per module (Staff, Machine), each
split visually by a "Positive/Negative" column with color coding, since both
modules together are still small enough to scan as one sheet apiece.
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent.parent
RESULTS_JSON = PROJECT_ROOT / 'test-report' / 'results' / 'results.json'
OUT_XLSX = PROJECT_ROOT / 'reports' / 'new-feature-test-report.xlsx'

if not RESULTS_JSON.exists():
    print(f'No results file at {RESULTS_JSON} - run the two new spec files with --reporter=json first.', file=sys.stderr)
    sys.exit(1)

OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

FONT = 'Arial'
NAVY = '1F3D2E'
TEAL = '157A5E'
PASS_GREEN = '1F9D63'
PASS_SOFT = 'E6F6EE'
FAIL_RED = 'D64545'
FAIL_SOFT = 'FDEAEA'
SKIP_AMBER = 'B98A2E'
SKIP_SOFT = 'FBF2E2'
NEG_PLUM = '7A3B69'
NEG_SOFT = 'F3E7F0'
POS_TEAL_SOFT = 'E7F3EE'
GREY = '5C6B63'
BORDER_CLR = 'DDE2DC'
thin = Side(style='thin', color=BORDER_CLR)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

ANSI_RE = re.compile(r'\x1b\[[0-9;]*m')
# "NEW-STF-006 (negative) - phone starts with 0 (0123456789) is rejected client-side"
TITLE_RE = re.compile(r'^(NEW-(?:STF|MCH)-\d+)\s*\(([a-zA-Z]+)\)\s*-\s*(.*)$')
STATUS_MAP = {'passed': 'PASS', 'failed': 'FAIL', 'timedOut': 'FAIL', 'interrupted': 'FAIL', 'skipped': 'SKIPPED'}


def walk_specs(suites, file_name=None):
    for suite in suites:
        fname = suite.get('file') or file_name
        for spec in suite.get('specs', []):
            yield fname, spec
        yield from walk_specs(suite.get('suites', []), fname)


data = json.loads(RESULTS_JSON.read_text(encoding='utf-8'))

rows = []
for fname, spec in walk_specs(data['suites']):
    match = TITLE_RE.match(spec['title'])
    if not match:
        continue
    tc_id, marker, scenario = match.group(1), match.group(2), match.group(3)
    category = marker.capitalize()

    test = spec['tests'][0]
    result = test['results'][0] if test['results'] else {}
    status = STATUS_MAP.get(result.get('status'), 'SKIPPED')
    duration_s = round((result.get('duration') or 0) / 1000, 1)

    error_message = ''
    if status == 'FAIL':
        errors = result.get('errors') or []
        msg = errors[0].get('message', '') if errors else 'Test failed.'
        error_message = ANSI_RE.sub('', msg).strip()

    if status == 'PASS':
        actual = 'All assertions passed as expected.'
    elif status == 'FAIL':
        actual = (error_message.split('\n')[0] or 'Test failed.')[:500]
    else:
        actual = 'Not executed (skipped).'

    module = 'Machine' if tc_id.startswith('NEW-MCH') else 'Staff'
    rows.append({
        'id': tc_id, 'module': module, 'category': category, 'scenario': scenario,
        'actual': actual, 'status': status, 'error_message': error_message,
        'duration': f'{duration_s}s',
    })

# Stable numeric order within each module.
rows.sort(key=lambda r: (r['module'], int(r['id'].rsplit('-', 1)[1])))

staff_rows = [r for r in rows if r['module'] == 'Staff']
machine_rows = [r for r in rows if r['module'] == 'Machine']

status_style = {
    'PASS': (PASS_GREEN, PASS_SOFT),
    'FAIL': (FAIL_RED, FAIL_SOFT),
    'SKIPPED': (SKIP_AMBER, SKIP_SOFT),
}
type_style = {
    'Positive': (TEAL, POS_TEAL_SOFT),
    'Negative': (NEG_PLUM, NEG_SOFT),
}
HEADERS = ['Test Case ID', 'Test Type', 'Scenario', 'Actual Result', 'Status', 'Error Message', 'Execution Time']
WIDTHS = {'A': 14, 'B': 11, 'C': 62, 'D': 40, 'E': 10, 'F': 40, 'G': 13}


def write_sheet(ws, sheet_rows, table_name):
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = border_all

    for r, row in enumerate(sheet_rows, start=2):
        values = [row['id'], row['category'], row['scenario'], row['actual'],
                  row['status'], row['error_message'], row['duration']]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=9.5)
            cell.border = border_all
            cell.alignment = Alignment(vertical='top', wrap_text=True)

        type_cell = ws.cell(row=r, column=2)
        tfg, tbg = type_style.get(row['category'], (GREY, 'FFFFFF'))
        type_cell.font = Font(name=FONT, size=9.5, bold=True, color=tfg)
        type_cell.fill = PatternFill('solid', fgColor=tbg)
        type_cell.alignment = Alignment(horizontal='center', vertical='center')

        status_cell = ws.cell(row=r, column=5)
        fg, bg = status_style.get(row['status'], (GREY, 'FFFFFF'))
        status_cell.font = Font(name=FONT, size=10, bold=True, color=fg)
        status_cell.fill = PatternFill('solid', fgColor=bg)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=r, column=1).font = Font(name=FONT, size=9.5, bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=r, column=7).alignment = Alignment(horizontal='center', vertical='top')
        ws.row_dimensions[r].height = 46

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    last_row = len(sheet_rows) + 1
    if sheet_rows:
        tab = Table(displayName=table_name, ref=f'A1:G{last_row}')
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showFirstColumn=False)
        ws.add_table(tab)


wb = Workbook()
wb.remove(wb.active)

ws_staff = wb.create_sheet('Staff Test Cases')
write_sheet(ws_staff, staff_rows, 'StaffTestCases')

ws_machine = wb.create_sheet('Machine Test Cases')
write_sheet(ws_machine, machine_rows, 'MachineTestCases')

# ---------- Summary sheet (first tab) ----------
ws2 = wb.create_sheet('Summary', 0)
ws2.sheet_view.showGridLines = False

ws2['B2'] = 'New Feature Coverage — E2E Test Report'
ws2['B2'].font = Font(name=FONT, size=16, bold=True, color=NAVY)
ws2['B3'] = 'Playwright · Chromium · Staff Management + Electric Idli Machine'
ws2['B3'].font = Font(name=FONT, size=11, italic=True, color=GREY)
ws2['B4'] = f'Run date: {data["stats"].get("startTime", "")[:10]}'
ws2['B4'].font = Font(name=FONT, size=10, color=GREY)


def count(sheet_rows, status):
    return sum(1 for r in sheet_rows if r['status'] == status)


row0 = 6
sections = [
    ('All Test Cases', rows),
    ('Staff Management', staff_rows),
    ('Electric Idli Machine', machine_rows),
]
r = row0
for label, sheet_rows in sections:
    ws2.cell(row=r, column=2, value=label).font = Font(name=FONT, size=11, bold=True, color=NAVY)
    r += 1
    for i, h in enumerate(['Total', 'Passed', 'Failed', 'Skipped', 'Positive', 'Negative'], start=2):
        c = ws2.cell(row=r, column=i, value=h)
        c.font = Font(name=FONT, size=9.5, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center')
    r += 1
    positive_n = sum(1 for x in sheet_rows if x['category'] == 'Positive')
    negative_n = sum(1 for x in sheet_rows if x['category'] == 'Negative')
    values = [len(sheet_rows), count(sheet_rows, 'PASS'), count(sheet_rows, 'FAIL'), count(sheet_rows, 'SKIPPED'), positive_n, negative_n]
    fills = [None, PASS_SOFT, FAIL_SOFT, SKIP_SOFT, POS_TEAL_SOFT, NEG_SOFT]
    fonts = [GREY, PASS_GREEN, FAIL_RED, SKIP_AMBER, TEAL, NEG_PLUM]
    for i, (val, fill, color) in enumerate(zip(values, fills, fonts), start=2):
        c = ws2.cell(row=r, column=i, value=val)
        c.font = Font(name=FONT, size=12, bold=True, color=color)
        c.alignment = Alignment(horizontal='center')
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
    r += 2

overall_row = r
total_failed = count(rows, 'FAIL')
total = len(rows)
total_passed = count(rows, 'PASS')
overall = 'FAIL' if total_failed > 0 else ('PASS' if total_passed == total else 'PARTIAL')
ws2.cell(row=overall_row, column=2, value='Overall Result').font = Font(name=FONT, size=11, bold=True, color=GREY)
oc = ws2.cell(row=overall_row, column=3, value=overall)
oc.font = Font(name=FONT, size=14, bold=True, color=TEAL if overall == 'PASS' else FAIL_RED)
ws2.cell(row=overall_row, column=2).fill = PatternFill('solid', fgColor='E7F3EE')
ws2.cell(row=overall_row, column=3).fill = PatternFill('solid', fgColor='E7F3EE')

note_row = overall_row + 2
ws2.cell(row=note_row, column=2, value='Sheets:').font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws2.cell(row=note_row + 1, column=2,
         value='"Staff Test Cases" and "Machine Test Cases" hold the full per-test detail (scenario, actual result, status, error message, execution time). This is a static export from one Playwright run - re-run the two spec files and this script to refresh it.').font = Font(name=FONT, size=10, color=GREY)
ws2.cell(row=note_row + 1, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws2.merge_cells(start_row=note_row + 1, start_column=2, end_row=note_row + 3, end_column=7)

for col, width in zip('ABCDEFG', [3, 24, 12, 10, 10, 10, 10]):
    ws2.column_dimensions[col].width = width

wb.save(str(OUT_XLSX))
print('Saved:', OUT_XLSX)
print(f'Staff:   {len(staff_rows)} (PASS={count(staff_rows, "PASS")} FAIL={count(staff_rows, "FAIL")} SKIPPED={count(staff_rows, "SKIPPED")})')
print(f'Machine: {len(machine_rows)} (PASS={count(machine_rows, "PASS")} FAIL={count(machine_rows, "FAIL")} SKIPPED={count(machine_rows, "SKIPPED")})')
