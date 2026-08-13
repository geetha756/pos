import json
from pathlib import Path
from openpyxl import Workbook

json_path = Path('playwright-test-output.json')
if not json_path.exists():
    raise SystemExit(f'Missing {json_path}')

# Playwright JSON output is UTF-16 encoded when redirected from the shell
text = json_path.read_text('utf-16')
report = json.loads(text)

specs = []
suites = report.get('suites', [])
if suites:
    first_suite = suites[0]
    nested_suites = first_suite.get('suites', [])
    if nested_suites:
        specs = nested_suites[0].get('specs', [])

if not specs:
    raise SystemExit('No spec entries found in Playwright JSON output')

test_rows = []
for idx, spec in enumerate(specs, start=1):
    title = spec.get('title', '')
    ok = spec.get('ok', False)
    status = 'PASSED' if ok else 'FAILED'
    location = spec.get('file', '')

    steps = f'Run Playwright test: {title}'
    expected = 'The application should satisfy the described staff management scenario.'
    actual = 'Passed' if status == 'PASSED' else 'See Playwright artifacts for details.'
    evidence = 'See playwright-report and test-results folders'

    test_rows.append({
        'Test Case ID': f'SM-{idx:02d}',
        'Scenario': title,
        'Steps': steps,
        'Expected Result': expected,
        'Actual Result': actual,
        'Pass/Fail': status,
        'Screenshot/Evidence': evidence,
        'Duration(ms)': spec.get('duration', 0),
        'Source': location,
    })

passed = sum(1 for row in test_rows if row['Pass/Fail'] == 'PASSED')
failed = sum(1 for row in test_rows if row['Pass/Fail'] != 'PASSED')
skipped = sum(1 for row in test_rows if row['Pass/Fail'] == 'SKIPPED')
total = len(test_rows)
overall = 'PASS' if failed == 0 else 'FAIL'

output_dir = Path('playwright-report/results')
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / 'StaffManagementPositiveFlowsReport.xlsx'

wb = Workbook()
ws = wb.active
ws.title = 'Staff Management Summary'

summary = [
    ('Total Tests', total),
    ('Passed', passed),
    ('Failed', failed),
    ('Skipped', skipped),
    ('Overall Status', overall),
]
for row_idx, (name, value) in enumerate(summary, start=1):
    ws.cell(row=row_idx, column=1, value=name)
    ws.cell(row=row_idx, column=2, value=value)

ws2 = wb.create_sheet(title='Test Details')
headers = ['Test Case ID', 'Scenario', 'Steps', 'Expected Result', 'Actual Result', 'Pass/Fail', 'Screenshot/Evidence', 'Duration(ms)', 'Source']
for col_idx, header in enumerate(headers, start=1):
    ws2.cell(row=1, column=col_idx, value=header)

for row_idx, row in enumerate(test_rows, start=2):
    for col_idx, header in enumerate(headers, start=1):
        ws2.cell(row=row_idx, column=col_idx, value=row[header])

wb.save(output_path)
print(f'Saved report to {output_path}')
print(f'Tests: total={total}, passed={passed}, failed={failed}, skipped={skipped}, overall={overall}')
