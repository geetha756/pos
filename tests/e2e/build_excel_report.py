"""
Builds the Staff Management E2E Excel report from a Playwright JSON reporter
run (test-report/results/results.json) plus the Test Case catalog
(tests/e2e/test-case-catalog.json). Run after `npm run test:e2e`:

    python tests/e2e/build_excel_report.py

Writes test-report/results/staff-management-e2e-report.xlsx. Run from
anywhere by passing the project root (the directory containing
playwright.config.js) as the first argument.
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent.parent
RESULTS_JSON = PROJECT_ROOT / 'test-report' / 'results' / 'results.json'
CATALOG_JSON = PROJECT_ROOT / 'tests' / 'e2e' / 'test-case-catalog.json'
OUT_XLSX = PROJECT_ROOT / 'test-report' / 'results' / 'staff-management-e2e-report.xlsx'

if not RESULTS_JSON.exists():
    print(f'No results file at {RESULTS_JSON} - run `npm run test:e2e` first.', file=sys.stderr)
    sys.exit(1)

FONT = 'Arial'
NAVY = '1F3D2E'
TEAL = '157A5E'
TEAL_SOFT = 'E7F3EE'
PASS_GREEN = '1F9D63'
PASS_SOFT = 'E6F6EE'
FAIL_RED = 'D64545'
FAIL_SOFT = 'FDEAEA'
SKIP_AMBER = 'B98A2E'
SKIP_SOFT = 'FBF2E2'
NEG_PLUM = '7A3B69'
NEG_SOFT = 'F3E7F0'
POS_TEAL_SOFT = 'E7F3EE'
GREY = '5C6B63'
BORDER_CLR = 'DDE2DC'
thin = Side(style='thin', color=BORDER_CLR)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

catalog = json.loads(CATALOG_JSON.read_text(encoding='utf-8'))
data = json.loads(RESULTS_JSON.read_text(encoding='utf-8'))

STATUS_MAP = {'passed': 'Pass', 'failed': 'Fail', 'timedOut': 'Fail', 'interrupted': 'Fail', 'skipped': 'Skipped'}


def walk_specs(suites, file_name=None):
    for suite in suites:
        fname = suite.get('file') or file_name
        for spec in suite.get('specs', []):
            yield fname, spec
        yield from walk_specs(suite.get('suites', []), fname)


rows = []
for fname, spec in walk_specs(data['suites']):
    title = spec['title']
    m = re.match(r'([A-Z]+-\d+):\s*(.*)', title)
    tc_id, scenario = (m.group(1), m.group(2)) if m else (title.split(':')[0], title)

    test = spec['tests'][0]
    result = test['results'][0] if test['results'] else {}
    status = STATUS_MAP.get(result.get('status'), 'Skipped')
    duration_s = round((result.get('duration') or 0) / 1000, 1)

    steps = [s['title'] for s in result.get('steps', []) if not s.get('title', '').startswith('expect.')]
    steps_text = '\n'.join(f'{i + 1}. {s}' for i, s in enumerate(steps))

    if status == 'Pass':
        actual = 'All assertions passed as expected.'
        evidence = 'See playwright-report/index.html (trace + video retained on failure only)'
    elif status == 'Fail':
        errors = result.get('errors') or []
        msg = errors[0].get('message', '') if errors else 'Test failed.'
        msg = re.sub(r'\x1b\[[0-9;]*m', '', msg)  # strip ANSI color codes
        actual = msg.strip().split('\n')[0][:500]
        shot = next((a['path'] for a in result.get('attachments', []) if a.get('name') == 'screenshot'), None)
        evidence = shot or 'See playwright-report/index.html for trace/video'
    else:
        actual = 'Not executed (skipped because a prior step in this test failed).'
        evidence = 'N/A - Skipped'

    cat = catalog.get(tc_id, {})
    category = cat.get('category') or ('Negative' if tc_id.startswith('NEG') else 'Positive')
    rows.append({
        'id': tc_id,
        'category': category,
        'scenario': cat.get('scenario', scenario),
        'preconditions': cat.get('preconditions', ''),
        'steps': steps_text,
        'expected': cat.get('expectedResult', ''),
        'actual': actual,
        'status': status,
        'evidence': evidence,
        'remarks': cat.get('remarks', ''),
        'file': fname,
        'duration': duration_s,
    })

wb = Workbook()

# ---------- Test Cases sheet ----------
ws = wb.active
ws.title = 'Test Cases'
ws.sheet_view.showGridLines = False

headers = ['Test Case ID', 'Test Type', 'Test Scenario', 'Preconditions', 'Test Steps', 'Expected Result',
           'Actual Result', 'Status', 'Screenshot / Evidence', 'Remarks']
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(vertical='center', wrap_text=True)
    c.border = border_all

status_style = {
    'Pass': (PASS_GREEN, PASS_SOFT),
    'Fail': (FAIL_RED, FAIL_SOFT),
    'Skipped': (SKIP_AMBER, SKIP_SOFT),
}
type_style = {
    'Positive': (TEAL, POS_TEAL_SOFT),
    'Negative': (NEG_PLUM, NEG_SOFT),
}

for r, row in enumerate(rows, start=2):
    values = [row['id'], row['category'], row['scenario'], row['preconditions'], row['steps'], row['expected'],
              row['actual'], row['status'], row['evidence'], row['remarks']]
    for c_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c_idx, value=val)
        cell.font = Font(name=FONT, size=9.5)
        cell.border = border_all
        cell.alignment = Alignment(vertical='top', wrap_text=True)

    type_cell = ws.cell(row=r, column=2)
    tfg, tbg = type_style.get(row['category'], (GREY, 'FFFFFF'))
    type_cell.font = Font(name=FONT, size=9.5, bold=True, color=tfg)
    type_cell.fill = PatternFill('solid', fgColor=tbg)
    type_cell.alignment = Alignment(horizontal='center', vertical='center')

    status_cell = ws.cell(row=r, column=8)
    fg, bg = status_style.get(row['status'], (GREY, 'FFFFFF'))
    status_cell.font = Font(name=FONT, size=10, bold=True, color=fg)
    status_cell.fill = PatternFill('solid', fgColor=bg)
    status_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=r, column=1).font = Font(name=FONT, size=9.5, bold=True)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='top')
    ws.row_dimensions[r].height = 90

widths = {'A': 12, 'B': 11, 'C': 34, 'D': 32, 'E': 40, 'F': 40, 'G': 30, 'H': 10, 'I': 26, 'J': 34}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = 'A2'
last_row = len(rows) + 1
tab = Table(displayName='TestCases', ref=f'A1:J{last_row}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showFirstColumn=False)
ws.add_table(tab)

# ---------- Summary sheet ----------
ws2 = wb.create_sheet('Summary', 0)
ws2.sheet_view.showGridLines = False

ws2['B2'] = 'Staff Management — E2E Test Report (Positive & Negative)'
ws2['B2'].font = Font(name=FONT, size=16, bold=True, color=NAVY)
ws2['B3'] = 'Playwright · Chromium · http://localhost:5000/staff/'
ws2['B3'].font = Font(name=FONT, size=11, italic=True, color=GREY)
ws2['B4'] = f'Run date: {data["stats"].get("startTime", "")[:10]}'
ws2['B4'].font = Font(name=FONT, size=10, color=GREY)

row0 = 6
metrics = [
    ('Total Tests', f'=COUNTA(\'Test Cases\'!A2:A{last_row})'),
    ('Passed', f'=COUNTIF(\'Test Cases\'!H2:H{last_row},"Pass")'),
    ('Failed', f'=COUNTIF(\'Test Cases\'!H2:H{last_row},"Fail")'),
    ('Skipped', f'=COUNTIF(\'Test Cases\'!H2:H{last_row},"Skipped")'),
    ('Pass Rate', ''),
]
for i, (label, formula) in enumerate(metrics):
    r = row0 + i
    ws2.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10, bold=True, color=GREY)
    if label == 'Pass Rate':
        formula = f'=C{row0 + 1}/C{row0}'
    c = ws2.cell(row=r, column=3, value=formula)
    c.font = Font(name=FONT, size=12, bold=True, color=TEAL)
    c.number_format = '0.0%' if label == 'Pass Rate' else '0'

ws2.cell(row=row0 + 1, column=2).fill = PatternFill('solid', fgColor=PASS_SOFT)
ws2.cell(row=row0 + 1, column=3).fill = PatternFill('solid', fgColor=PASS_SOFT)
ws2.cell(row=row0 + 1, column=3).font = Font(name=FONT, size=12, bold=True, color=PASS_GREEN)
ws2.cell(row=row0 + 2, column=2).fill = PatternFill('solid', fgColor=FAIL_SOFT)
ws2.cell(row=row0 + 2, column=3).fill = PatternFill('solid', fgColor=FAIL_SOFT)
ws2.cell(row=row0 + 2, column=3).font = Font(name=FONT, size=12, bold=True, color=FAIL_RED)
ws2.cell(row=row0 + 3, column=2).fill = PatternFill('solid', fgColor=SKIP_SOFT)
ws2.cell(row=row0 + 3, column=3).fill = PatternFill('solid', fgColor=SKIP_SOFT)
ws2.cell(row=row0 + 3, column=3).font = Font(name=FONT, size=12, bold=True, color=SKIP_AMBER)

overall_row = row0 + len(metrics) + 1
ws2.cell(row=overall_row, column=2, value='Overall Result').font = Font(name=FONT, size=11, bold=True, color=GREY)
overall_formula = f'=IF(C{row0 + 2}>0,"FAIL",IF(C{row0}=C{row0 + 1},"PASS","PARTIAL"))'
oc = ws2.cell(row=overall_row, column=3, value=overall_formula)
oc.font = Font(name=FONT, size=14, bold=True, color=TEAL)
ws2.cell(row=overall_row, column=2).fill = PatternFill('solid', fgColor=TEAL_SOFT)
ws2.cell(row=overall_row, column=3).fill = PatternFill('solid', fgColor=TEAL_SOFT)

# --- Positive vs Negative breakdown ---
breakdown_row = overall_row + 2
ws2.cell(row=breakdown_row, column=2, value='By Test Type').font = Font(name=FONT, size=11, bold=True, color=NAVY)

bh_row = breakdown_row + 1
for i, h in enumerate(['Type', 'Total', 'Passed', 'Failed'], start=2):
    c = ws2.cell(row=bh_row, column=i, value=h)
    c.font = Font(name=FONT, size=9.5, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center')

for i, ttype in enumerate(['Positive', 'Negative']):
    r = bh_row + 1 + i
    ws2.cell(row=r, column=2, value=ttype).font = Font(name=FONT, size=10, bold=True,
                                                        color=TEAL if ttype == 'Positive' else NEG_PLUM)
    ws2.cell(row=r, column=3, value=f'=COUNTIF(\'Test Cases\'!B2:B{last_row},"{ttype}")').number_format = '0'
    ws2.cell(row=r, column=4, value=(
        f'=COUNTIFS(\'Test Cases\'!B2:B{last_row},"{ttype}",\'Test Cases\'!H2:H{last_row},"Pass")'
    )).number_format = '0'
    ws2.cell(row=r, column=5, value=(
        f'=COUNTIFS(\'Test Cases\'!B2:B{last_row},"{ttype}",\'Test Cases\'!H2:H{last_row},"Fail")'
    )).number_format = '0'
    for col in range(3, 6):
        ws2.cell(row=r, column=col).font = Font(name=FONT, size=10)
        ws2.cell(row=r, column=col).alignment = Alignment(horizontal='center')

note_row = bh_row + 4
ws2.cell(row=note_row, column=2, value='Finding (minor, non-blocking):').font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws2.cell(row=note_row + 1, column=2, value=(
    'staffForm on Add/Edit Staff Member has no novalidate attribute, so the browser\'s native constraint '
    'validation intercepts before the page\'s submit handler runs - Bootstrap\'s custom "Phone number must be '
    'exactly 10 digits." message never actually renders on a real submit attempt (the browser\'s own generic '
    'validation tooltip shows instead). Invalid submissions are still correctly blocked either way (see TC-008). '
    'Not fixed - application code was left untouched per instructions.'
))
ws2.cell(row=note_row + 1, column=2).font = Font(name=FONT, size=10, color=GREY)
ws2.cell(row=note_row + 1, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws2.merge_cells(start_row=note_row + 1, start_column=2, end_row=note_row + 4, end_column=7)

ws2.cell(row=note_row + 6, column=2, value='Reports:').font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws2.cell(row=note_row + 7, column=2, value='playwright-report/index.html and test-report/results/ (this workbook + HTML report + raw JSON results)').font = Font(name=FONT, size=10, color=GREY)

for col, width in zip('ABCDEFG', [3, 24, 14, 10, 10, 3, 3]):
    ws2.column_dimensions[col].width = width

wb.save(str(OUT_XLSX))
print('Saved:', OUT_XLSX)
print('Rows:', len(rows))
for r in rows:
    print(' -', r['id'], f"[{r['category']}]", r['status'], r['scenario'][:55])
