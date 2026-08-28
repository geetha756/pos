"""
Builds reports/staff-test-report.xlsx from a Playwright JSON reporter run
(test-report/results/results.json) plus the Test Case catalog
(tests/e2e/test-case-catalog.json). Run after the e2e suite:

    npx playwright test
    python tests/e2e/build_staff_test_report.py

Workbook layout: Summary, Positive Test Cases, Negative Test Cases - split
into separate sheets (rather than one combined sheet with a Type column) so
positive and negative coverage are each reviewable on their own.
Columns: Test Case ID, Test Type, Test Scenario, Expected Result, Actual
Result, Status, Error Message, Execution Time, Screenshot Path.

Run from anywhere by passing the project root (the directory containing
playwright.config.js) as the first argument.
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent.parent
RESULTS_JSON = PROJECT_ROOT / 'test-report' / 'results' / 'results.json'
CATALOG_JSON = PROJECT_ROOT / 'tests' / 'e2e' / 'test-case-catalog.json'
OUT_XLSX = PROJECT_ROOT / 'reports' / 'staff-test-report.xlsx'

if not RESULTS_JSON.exists():
    print(f'No results file at {RESULTS_JSON} - run `npx playwright test` first.', file=sys.stderr)
    sys.exit(1)

OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

FONT = 'Arial'
NAVY = '1F3D2E'
TEAL = '157A5E'
PASS_GREEN = '1F9D63'
PASS_SOFT = 'E6F6EE'
FAIL_RED = 'D64545'
FAIL_SOFT = 'FDEAEA'
SKIP_AMBER = 'B98A2E'
SKIP_SOFT = 'FBF2E2'
NEG_PLUM = '7A3B69'
NEG_SOFT = 'F3E7F0'
POS_TEAL_SOFT = 'E7F3EE'
GREY = '5C6B63'
BORDER_CLR = 'DDE2DC'
thin = Side(style='thin', color=BORDER_CLR)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

catalog = json.loads(CATALOG_JSON.read_text(encoding='utf-8')) if CATALOG_JSON.exists() else {}
data = json.loads(RESULTS_JSON.read_text(encoding='utf-8'))

STATUS_MAP = {'passed': 'PASS', 'failed': 'FAIL', 'timedOut': 'FAIL', 'interrupted': 'FAIL', 'skipped': 'SKIPPED'}
ANSI_RE = re.compile(r'\x1b\[[0-9;]*m')


def walk_specs(suites, file_name=None):
    for suite in suites:
        fname = suite.get('file') or file_name
        for spec in suite.get('specs', []):
            yield fname, spec
        yield from walk_specs(suite.get('suites', []), fname)


rows = []
for fname, spec in walk_specs(data['suites']):
    title = spec['title']
    m = re.match(r'([A-Z]+-\d+):\s*(.*)', title)
    tc_id, scenario_from_title = (m.group(1), m.group(2)) if m else (title.split(':')[0], title)

    test = spec['tests'][0]
    result = test['results'][0] if test['results'] else {}
    status = STATUS_MAP.get(result.get('status'), 'SKIPPED')
    duration_s = round((result.get('duration') or 0) / 1000, 1)

    error_message = ''
    if status == 'FAIL':
        errors = result.get('errors') or []
        msg = errors[0].get('message', '') if errors else 'Test failed.'
        error_message = ANSI_RE.sub('', msg).strip()

    if status == 'PASS':
        actual = 'All assertions passed as expected.'
        screenshot_path = 'N/A - test passed (screenshots are only captured on failure)'
    elif status == 'FAIL':
        actual = (error_message.split('\n')[0] or 'Test failed.')[:500]
        shot = next((a['path'] for a in result.get('attachments', []) if a.get('name') == 'screenshot'), None)
        screenshot_path = shot or 'See playwright-report/index.html for trace/video'
    else:
        actual = 'Not executed (skipped because a prior step in this test failed).'
        screenshot_path = 'N/A - skipped'

    cat = catalog.get(tc_id, {})
    category = cat.get('category') or ('Negative' if tc_id.startswith('NEG') else 'Positive')
    rows.append({
        'id': tc_id,
        'category': category,
        'scenario': cat.get('scenario', scenario_from_title),
        'expected': cat.get('expectedResult', ''),
        'actual': actual,
        'status': status,
        'error_message': error_message,
        'duration': f'{duration_s}s',
        'screenshot_path': screenshot_path,
    })

positive_rows = [r for r in rows if r['category'] == 'Positive']
negative_rows = [r for r in rows if r['category'] == 'Negative']

status_style = {
    'PASS': (PASS_GREEN, PASS_SOFT),
    'FAIL': (FAIL_RED, FAIL_SOFT),
    'SKIPPED': (SKIP_AMBER, SKIP_SOFT),
}
type_style = {
    'Positive': (TEAL, POS_TEAL_SOFT),
    'Negative': (NEG_PLUM, NEG_SOFT),
}
HEADERS = ['Test Case ID', 'Test Type', 'Test Scenario', 'Expected Result', 'Actual Result',
           'Status', 'Error Message', 'Execution Time', 'Screenshot Path']
WIDTHS = {'A': 12, 'B': 11, 'C': 40, 'D': 46, 'E': 34, 'F': 10, 'G': 40, 'H': 13, 'I': 34}


def write_sheet(ws, sheet_rows, table_name):
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = border_all

    for r, row in enumerate(sheet_rows, start=2):
        values = [row['id'], row['category'], row['scenario'], row['expected'], row['actual'],
                  row['status'], row['error_message'], row['duration'], row['screenshot_path']]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=9.5)
            cell.border = border_all
            cell.alignment = Alignment(vertical='top', wrap_text=True)

        type_cell = ws.cell(row=r, column=2)
        tfg, tbg = type_style.get(row['category'], (GREY, 'FFFFFF'))
        type_cell.font = Font(name=FONT, size=9.5, bold=True, color=tfg)
        type_cell.fill = PatternFill('solid', fgColor=tbg)
        type_cell.alignment = Alignment(horizontal='center', vertical='center')

        status_cell = ws.cell(row=r, column=6)
        fg, bg = status_style.get(row['status'], (GREY, 'FFFFFF'))
        status_cell.font = Font(name=FONT, size=10, bold=True, color=fg)
        status_cell.fill = PatternFill('solid', fgColor=bg)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=r, column=1).font = Font(name=FONT, size=9.5, bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center', vertical='top')
        ws.row_dimensions[r].height = 60

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    last_row = len(sheet_rows) + 1
    if sheet_rows:
        tab = Table(displayName=table_name, ref=f'A1:I{last_row}')
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showFirstColumn=False)
        ws.add_table(tab)


wb = Workbook()
wb.remove(wb.active)

ws_pos = wb.create_sheet('Positive Test Cases')
write_sheet(ws_pos, positive_rows, 'PositiveTestCases')

ws_neg = wb.create_sheet('Negative Test Cases')
write_sheet(ws_neg, negative_rows, 'NegativeTestCases')

# ---------- Summary sheet (first tab) ----------
ws2 = wb.create_sheet('Summary', 0)
ws2.sheet_view.showGridLines = False

ws2['B2'] = 'Staff Management — E2E Test Report'
ws2['B2'].font = Font(name=FONT, size=16, bold=True, color=NAVY)
ws2['B3'] = 'Playwright · Chromium · http://localhost:5000/staff/'
ws2['B3'].font = Font(name=FONT, size=11, italic=True, color=GREY)
ws2['B4'] = f'Run date: {data["stats"].get("startTime", "")[:10]}'
ws2['B4'].font = Font(name=FONT, size=10, color=GREY)


def count(sheet_rows, status):
    return sum(1 for r in sheet_rows if r['status'] == status)


row0 = 6
sections = [
    ('All Test Cases', rows),
    ('Positive Test Cases', positive_rows),
    ('Negative Test Cases', negative_rows),
]
r = row0
for label, sheet_rows in sections:
    ws2.cell(row=r, column=2, value=label).font = Font(name=FONT, size=11, bold=True, color=NAVY)
    r += 1
    for i, h in enumerate(['Total', 'Passed', 'Failed', 'Skipped'], start=2):
        c = ws2.cell(row=r, column=i, value=h)
        c.font = Font(name=FONT, size=9.5, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center')
    r += 1
    values = [len(sheet_rows), count(sheet_rows, 'PASS'), count(sheet_rows, 'FAIL'), count(sheet_rows, 'SKIPPED')]
    fills = [None, PASS_SOFT, FAIL_SOFT, SKIP_SOFT]
    fonts = [GREY, PASS_GREEN, FAIL_RED, SKIP_AMBER]
    for i, (val, fill, color) in enumerate(zip(values, fills, fonts), start=2):
        c = ws2.cell(row=r, column=i, value=val)
        c.font = Font(name=FONT, size=12, bold=True, color=color)
        c.alignment = Alignment(horizontal='center')
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
    r += 2

overall_row = r
total_failed = count(rows, 'FAIL')
total = len(rows)
total_passed = count(rows, 'PASS')
overall = 'FAIL' if total_failed > 0 else ('PASS' if total_passed == total else 'PARTIAL')
ws2.cell(row=overall_row, column=2, value='Overall Result').font = Font(name=FONT, size=11, bold=True, color=GREY)
oc = ws2.cell(row=overall_row, column=3, value=overall)
oc.font = Font(name=FONT, size=14, bold=True, color=TEAL)
ws2.cell(row=overall_row, column=2).fill = PatternFill('solid', fgColor='E7F3EE')
ws2.cell(row=overall_row, column=3).fill = PatternFill('solid', fgColor='E7F3EE')

note_row = overall_row + 2
ws2.cell(row=note_row, column=2, value='Sheets:').font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws2.cell(row=note_row + 1, column=2,
         value='"Positive Test Cases" and "Negative Test Cases" hold the full per-test detail (scenario, expected/actual result, status, error message, execution time, screenshot path). This is a static export - re-run tests and this script to refresh it.').font = Font(name=FONT, size=10, color=GREY)
ws2.cell(row=note_row + 1, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws2.merge_cells(start_row=note_row + 1, start_column=2, end_row=note_row + 3, end_column=7)

for col, width in zip('ABCDEFG', [3, 24, 14, 10, 10, 3, 3]):
    ws2.column_dimensions[col].width = width

wb.save(str(OUT_XLSX))
print('Saved:', OUT_XLSX)
print(f'Positive: {len(positive_rows)} (PASS={count(positive_rows, "PASS")} FAIL={count(positive_rows, "FAIL")} SKIPPED={count(positive_rows, "SKIPPED")})')
print(f'Negative: {len(negative_rows)} (PASS={count(negative_rows, "PASS")} FAIL={count(negative_rows, "FAIL")} SKIPPED={count(negative_rows, "SKIPPED")})')
